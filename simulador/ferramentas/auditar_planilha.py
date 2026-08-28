#!/usr/bin/env python3
"""Despeja a estrutura completa de uma pasta de trabalho .xlsx.

Lê a planilha duas vezes: uma com as fórmulas e outra com os valores em cache
gravados pelo Excel. Cada célula sai com as duas faces lado a lado, que é o que
permite conferir o que a fórmula faz contra o número que o Excel produziu.

Uso:
    auditar_planilha.py <arquivo.xlsx> <diretorio_de_saida>

Gera, no diretório de saída:
    inventario.json     resumo de abas, nomes definidos, validações, mesclagens
    celulas.tsv         uma linha por célula não vazia, com fórmula e valor
    formulas.txt        só as fórmulas, agrupadas por aba, para leitura corrida
"""

import json
import os
import sys
import zipfile

import openpyxl
from openpyxl.utils import get_column_letter


def erros_excel(valor):
    """Os códigos de erro que o item 39 do prompt manda registrar como ABERTO."""
    if not isinstance(valor, str):
        return None
    for código in ("#REF!", "#VALUE!", "#N/A", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!"):
        if código in valor:
            return código
    return None


def inventariar(caminho, destino):
    os.makedirs(destino, exist_ok=True)

    wb_f = openpyxl.load_workbook(caminho, data_only=False)
    wb_v = openpyxl.load_workbook(caminho, data_only=True)

    inventario = {
        "arquivo": os.path.basename(caminho),
        "bytes": os.path.getsize(caminho),
        "abas": [],
        "nomes_definidos": [],
        "ocorrencias": [],
    }

    for nome, dn in wb_f.defined_names.items():
        inventario["nomes_definidos"].append({"nome": nome, "referencia": str(dn.value)})

    linhas = ["nome_aba\tcelula\ttipo\tformula\tvalor_cache\tformato_numero"]
    blocos_formula = []

    for aba in wb_f.sheetnames:
        ws_f = wb_f[aba]
        ws_v = wb_v[aba]

        resumo = {
            "nome": aba,
            "visibilidade": ws_f.sheet_state,
            "dimensao": ws_f.dimensions,
            "max_linha": ws_f.max_row,
            "max_coluna": ws_f.max_column,
            "mescladas": [str(m) for m in ws_f.merged_cells.ranges],
            "validacoes": [],
            "formatacao_condicional": [],
            "celulas_preenchidas": 0,
            "celulas_formula": 0,
        }

        for dv in ws_f.data_validations.dataValidation:
            resumo["validacoes"].append(
                {
                    "tipo": dv.type,
                    "formula1": dv.formula1,
                    "formula2": dv.formula2,
                    "aplica_em": str(dv.sqref),
                }
            )

        try:
            for intervalo in ws_f.conditional_formatting:
                resumo["formatacao_condicional"].append(
                    {
                        "intervalo": str(intervalo.sqref),
                        "regras": [str(r.type) for r in intervalo.rules],
                    }
                )
        except Exception as exc:  # formatação exótica não deve derrubar a auditoria
            resumo["formatacao_condicional"].append({"erro": repr(exc)})

        formulas_da_aba = []

        for linha_f, linha_v in zip(ws_f.iter_rows(), ws_v.iter_rows()):
            for cel_f, cel_v in zip(linha_f, linha_v):
                if cel_f.value is None and cel_v.value is None:
                    continue
                resumo["celulas_preenchidas"] += 1

                é_formula = isinstance(cel_f.value, str) and cel_f.value.startswith("=")
                formula = cel_f.value if é_formula else ""
                if é_formula:
                    resumo["celulas_formula"] += 1
                    formulas_da_aba.append(f"{cel_f.coordinate}\t{cel_f.value}")

                valor = cel_v.value if é_formula else cel_f.value

                código = erros_excel(valor) or erros_excel(formula)
                if código:
                    inventario["ocorrencias"].append(
                        {
                            "aba": aba,
                            "celula": cel_f.coordinate,
                            "erro": código,
                            "formula": formula,
                            "valor": str(valor),
                        }
                    )

                def limpar(x):
                    return str(x).replace("\t", " ").replace("\n", " \\n ")

                linhas.append(
                    "\t".join(
                        [
                            aba,
                            cel_f.coordinate,
                            "formula" if é_formula else type(cel_f.value).__name__,
                            limpar(formula),
                            limpar(valor) if valor is not None else "",
                            str(cel_f.number_format),
                        ]
                    )
                )

        inventario["abas"].append(resumo)
        if formulas_da_aba:
            blocos_formula.append(f"### ABA: {aba}\n" + "\n".join(formulas_da_aba))

    with open(os.path.join(destino, "inventario.json"), "w", encoding="utf-8") as fh:
        json.dump(inventario, fh, ensure_ascii=False, indent=2)
    with open(os.path.join(destino, "celulas.tsv"), "w", encoding="utf-8") as fh:
        fh.write("\n".join(linhas))
    with open(os.path.join(destino, "formulas.txt"), "w", encoding="utf-8") as fh:
        fh.write("\n\n".join(blocos_formula))

    # A lista de partes do zip revela gráficos, tabelas e vínculos externos que
    # a openpyxl não expõe diretamente.
    with zipfile.ZipFile(caminho) as z:
        partes = sorted(n for n in z.namelist() if "chart" in n or "table" in n or "external" in n)

    print(f"abas: {len(inventario['abas'])}")
    for a in inventario["abas"]:
        print(
            f"  {a['nome']!r:40} {a['visibilidade']:8} {a['dimensao']:15}"
            f" preenchidas={a['celulas_preenchidas']:6} formulas={a['celulas_formula']:6}"
            f" validacoes={len(a['validacoes'])}"
        )
    print(f"nomes definidos: {len(inventario['nomes_definidos'])}")
    print(f"celulas com erro: {len(inventario['ocorrencias'])}")
    print(f"partes auxiliares no zip: {len(partes)}")
    for p in partes:
        print(f"  {p}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        sys.exit(__doc__)
    inventariar(sys.argv[1], sys.argv[2])
