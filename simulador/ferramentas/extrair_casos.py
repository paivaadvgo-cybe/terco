#!/usr/bin/env python3
"""Extrai da planilha os casos de equivalência, célula a célula.

Cada aba de produto guarda um estado salvo: uma operação que alguém simulou, e
o cronograma que o Excel calculou para ela. É esse par — entrada e saída — que
serve de caso de teste, e ele vale mais que qualquer número transcrito à mão:
sai do próprio arquivo, com todos os dígitos que o Excel gravou.

O que se extrai por aba:
    entrada     valor, prazo, carência, taxas, e o valor financiado em cache
    encargos    TAC, IOF, garantias, renda para aval, alienação
    cronograma  cada parcela, com juros, amortização, saldo e prestação
    tir         as duas taxas internas de retorno

Onde a planilha não tem valor — porque a fórmula não foi arrastada até ali —
a parcela simplesmente não entra. O que ela não calculou não pode ser
comparado, e fingir que zero é resposta seria pior que não comparar.

Uso:
    extrair_casos.py <arquivo.xlsx> <saida.json>
"""

import json
import sys

import openpyxl

# Onde as abas mensais guardam as escolhas de cobrança.
BANDEIRAS_MENSAIS = {
    'tacFinanciada': ('AA1', 2),      # 1 descontada, 2 financiada
    'iofFinanciado': ('AB1', 2),
}

# O bloco de IOF das abas mensais. `normalTruncado` e `simplesTruncado` são as
# somas que a planilha faz sobre faixas fixas de linhas — é aí que ABERTO-08
# aparece, e é por isso que elas são extraídas separadas do total.
BLOCO_IOF_MENSAL = {
    'incide': 'L19',
    'adicional': 'AE24',
    'normalTruncado': 'AE25',
    'simplesTruncado': 'AE26',
    'total': 'AF24',
    'aplicado': 'F16',
}


# Como cada aba escreve a mesma operação. Descoberto pela auditoria; as colunas
# do cronograma mudam entre as famílias, e é por isso que elas são declaradas
# em vez de deduzidas.
ABAS = {
    'Linhas Investimento': {
        'produto': 'investimento', 'colunas': 'ABCDEF',
        'taxa': ('C23', 'mensal'), 'convencao': 'mensalComposta',
        'base': 'planilha', 'carencia': 'pagos',
        'encargos': {'tac': 'F15', 'iof': 'F16', 'garantia': 'F17',
                     'rendaParaAval': 'B19', 'alienacao': 'D19'},
        'bandeiras': BANDEIRAS_MENSAIS, 'iof': BLOCO_IOF_MENSAL,
        'tir': {'comBonus': 'AJ22', 'semBonus': 'AK22'},
    },
    'Linhas Giro Puro': {
        'produto': 'giro', 'colunas': 'ABCDEF',
        'taxa': ('C23', 'mensal'), 'convencao': 'mensalComposta',
        'base': 'valorFinanciado', 'carencia': 'pagos',
        'encargos': {'tac': 'F15', 'iof': 'F16', 'garantia': 'F17',
                     'rendaParaAval': 'B19', 'alienacao': 'D19'},
        'bandeiras': BANDEIRAS_MENSAIS, 'iof': BLOCO_IOF_MENSAL,
        'tir': {'comBonus': 'AJ22', 'semBonus': 'AK22'},
    },
    'Linhas Transportes': {
        'produto': 'transportes', 'colunas': 'ABCDEF',
        'taxa': ('C23', 'mensal'), 'convencao': 'mensalComposta',
        'base': 'planilha', 'carencia': 'pagos',
        'encargos': {'tac': 'F15', 'iof': 'F16', 'garantia': 'F17',
                     'rendaParaAval': 'B19', 'alienacao': 'D19'},
        'bandeiras': BANDEIRAS_MENSAIS, 'iof': BLOCO_IOF_MENSAL,
        'tir': {},
    },
    'Mais Crédito': {
        'produto': 'microcredito', 'colunas': 'ABCDEF',
        'taxa': ('C23', 'mensal'), 'convencao': 'mensalComposta',
        'base': 'planilha', 'carencia': 'pagos',
        'encargos': {'tac': 'F15', 'iof': 'F16', 'garantia': 'F17',
                     'rendaParaAval': 'B19'},
        'bandeiras': BANDEIRAS_MENSAIS, 'iof': BLOCO_IOF_MENSAL,
        'tir': {},
    },
    'Mais Crédito (2)': {
        'produto': 'microcredito', 'colunas': 'ABCDEF',
        'taxa': ('C23', 'mensal'), 'convencao': 'mensalComposta',
        'base': 'planilha', 'carencia': 'pagos',
        'encargos': {'tac': 'F15', 'iof': 'F16', 'garantia': 'F17',
                     'rendaParaAval': 'B19'},
        'bandeiras': BANDEIRAS_MENSAIS, 'iof': BLOCO_IOF_MENSAL,
        'tir': {},
    },
    # Fungetur e FINEP têm uma coluna a mais: o juro do indexador fica em D, e
    # a amortização escorrega para E, o saldo para F e a prestação para G.
    'Linhas Fungetur': {
        'produto': 'fungetur', 'colunas': 'ABCEFG', 'colunaIndexador': 'D',
        'taxa': ('C23', 'anual'), 'indexador': ('D23', 'anual'),
        'convencao': 'diasUteis', 'base': 'valorFinanciado', 'carencia': 'pagos',
        'financiadoEm': 'F23',
        # A célula G16 destas abas não é IOF: a fórmula é a do encargo do
        # FAMPE, e o rótulo ao lado diz "CCA FAMPE". O IOF delas vive noutro
        # bloco, e não é comparado aqui.
        'encargos': {'tac': 'G15', 'garantia': 'G16'},
        'tir': {},
    },
    'Linhas FINEP': {
        'produto': 'finep', 'colunas': 'ABCEFG', 'colunaIndexador': 'D',
        'taxa': ('D15', 'anual'), 'indexador': ('D16', 'anual'),
        'convencao': 'diasUteis', 'base': 'valorFinanciado', 'carencia': 'pagos',
        'financiadoEm': 'F23',
        'encargos': {'tac': 'G15', 'iof': 'G16'},
        'tir': {},
    },
    # A prestação destas duas abas é a coluna G, não a F. A F soma juros e
    # amortização sem a guarda da carência, e na carência capitalizada do
    # Produtor Empreendedor ela mostra um pagamento que não acontece — R$ 297,14
    # na parcela 18, onde nada é pago. Quem tem razão é a G, e é a G que a
    # própria planilha lê para calcular a renda exigida do avalista.
    'FCO Empresarial': {
        'produto': 'fco', 'colunas': 'ABCDEG',
        'taxa': ('D16', 'anual'), 'convencao': 'mensalComposta',
        'base': 'planilha', 'carencia': 'pagos',
        'encargos': {'tac': 'G15'},
        'tir': {},
    },
    # A linha 2 do Produtor Empreendedor capitaliza os juros da carência, e a
    # amortização divide o saldo ao fim dela, não o valor financiado.
    'Produtor Empreendedor': {
        'produto': 'produtorEmpreendedor', 'colunas': 'ABCDEG',
        'taxa': ('C23', 'mensal'), 'convencao': 'mensalComposta',
        'base': 'valorFinanciado', 'carencia': 'capitalizados',
        'financiadoEm': 'BG23',
        'encargos': {'tac': 'G15', 'garantia': 'E15'},
        'tir': {},
    },
}

PRIMEIRA_LINHA = 24     # a parcela 1; a 23 é a parcela zero, só com o saldo


def numero(valor):
    return float(valor) if isinstance(valor, (int, float)) else None


def extrair(caminho, saida):
    wb = openpyxl.load_workbook(caminho, data_only=True)
    casos = []

    for aba, cfg in ABAS.items():
        if aba not in wb.sheetnames:
            continue
        ws = wb[aba]
        cols = cfg['colunas']
        prazo = numero(ws['B17'].value)
        if prazo is None:
            continue

        taxa = numero(ws[cfg['taxa'][0]].value)
        financiado = numero(ws[cfg.get('financiadoEm', 'E23')].value)

        parcelas = []
        for n in range(1, int(prazo) + 1):
            linha = PRIMEIRA_LINHA + n - 1
            juros = numero(ws[f'{cols[2]}{linha}'].value)
            amortizacao = numero(ws[f'{cols[3]}{linha}'].value)
            saldo = numero(ws[f'{cols[4]}{linha}'].value)
            prestacao = numero(ws[f'{cols[5]}{linha}'].value)
            # Onde a fórmula não foi arrastada, não há o que comparar.
            if None in (juros, amortizacao, saldo, prestacao):
                break
            registro = {
                'parcela': n, 'juros': juros, 'amortizacao': amortizacao,
                'saldo': saldo, 'prestacao': prestacao,
            }
            if 'colunaIndexador' in cfg:
                registro['jurosIndexador'] = numero(ws[f"{cfg['colunaIndexador']}{linha}"].value)
            parcelas.append(registro)

        casos.append({
            'aba': aba,
            'produto': cfg['produto'],
            'entrada': {
                'valorSolicitado': numero(ws['B15'].value),
                'prazo': int(prazo),
                'carencia': int(numero(ws['B16'].value) or 0),
                'valorFinanciado': financiado,
                'taxa': {'valor': taxa, 'unidade': cfg['taxa'][1]},
                'indexador': ({'valor': numero(ws[cfg['indexador'][0]].value),
                               'unidade': cfg['indexador'][1]} if 'indexador' in cfg else None),
                'convencaoTaxa': cfg['convencao'],
                'baseAmortizacao': cfg['base'],
                'tratamentoCarencia': cfg['carencia'],
                'dataProposta': str(ws['D17'].value)[:10] if ws['D17'].value else '2026-08-06',
            },
            'bandeiras': {
                nome: ws[celula].value == ligado
                for nome, (celula, ligado) in cfg.get('bandeiras', {}).items()
            },
            'iof': ({
                nome: (ws[celula].value is True if nome == 'incide' else numero(ws[celula].value))
                for nome, celula in cfg['iof'].items()
            } if 'iof' in cfg else None),
            'encargos': {nome: numero(ws[celula].value) for nome, celula in cfg['encargos'].items()},
            'tir': {nome: numero(ws[celula].value) for nome, celula in cfg['tir'].items()},
            'parcelasComValor': len(parcelas),
            'parcelas': parcelas,
        })

    with open(saida, 'w', encoding='utf-8') as fh:
        json.dump({'origem': caminho.split('/')[-1], 'casos': casos}, fh, ensure_ascii=False, indent=1)

    print(f'casos extraídos: {len(casos)}')
    for c in casos:
        completo = c['parcelasComValor'] == c['entrada']['prazo']
        print(f"  {c['aba']:26} prazo {c['entrada']['prazo']:3} · "
              f"{c['parcelasComValor']:3} parcelas com valor{'' if completo else '  (truncado)'}")


if __name__ == '__main__':
    if len(sys.argv) != 3:
        sys.exit(__doc__)
    extrair(sys.argv[1], sys.argv[2])
