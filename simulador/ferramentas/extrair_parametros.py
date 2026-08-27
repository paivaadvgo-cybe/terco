#!/usr/bin/env python3
"""Extrai da planilha os parâmetros financeiros, em JSON estruturado.

A planilha guarda as mesmas grandezas em unidades diferentes conforme a aba:
'Linhas Giro Puro' escreve as taxas em pontos percentuais (2,27), as demais
abas mensais em decimal (0,0217), e FCO, Fungetur e FINEP em taxa anual — o
FCO em decimal e as outras duas em pontos percentuais. Nada aqui converte
nada: cada taxa sai acompanhada da unidade em que foi encontrada, e a
conversão é decisão do motor. Perder essa distinção é como se troca 2,17%
por 0,0217 sem perceber.

Uso:
    extrair_parametros.py <arquivo.xlsx> <saida.json>
"""

import json
import os
import re
import sys

import openpyxl

# Onde ficam as tabelas de linhas de crédito de cada aba, e em que unidade as
# taxas estão escritas. Descoberto pela auditoria; ver MAPA_DE_REGRAS.
LAYOUTS = {
    "mensal_decimal": {
        "cabecalho": 4,
        "colunas": {
            "nome": "N", "taxaCheia": "O", "taxaBonus": "P",
            "taxaCG1": "Q", "taxaCG2": "R",
            "prazoMaximo": "S", "carenciaMaxima": "T", "limite": "U",
        },
        "unidade": "mensal",
        "escala": 1.0,
    },
    "mensal_percentual": {          # anomalia isolada de 'Linhas Giro Puro'
        "cabecalho": 4,
        "colunas": {
            "nome": "N", "taxaCheia": "O", "taxaBonus": "P",
            "taxaCG1": "Q", "taxaCG2": "R",
            "prazoMaximo": "S", "carenciaMaxima": "T", "limite": "U",
        },
        "unidade": "mensal",
        "escala": 0.01,
    },
    "anual_decimal": {              # FCO: rótulo diz "a.m.", os valores são anuais
        "cabecalho": 4,
        "colunas": {
            "nome": "N", "taxaCheia": "O", "taxaBonus": "P",
            "prazoMaximo": "S", "carenciaMaxima": "T", "limite": "U",
        },
        "unidade": "anual",
        "escala": 1.0,
    },
    "anual_percentual": {           # Fungetur
        "cabecalho": 3,
        "colunas": {
            "nome": "N", "taxaCheia": "O",
            "prazoMaximo": "Q", "carenciaMaxima": "R", "limite": "S",
        },
        "unidade": "anual",
        "escala": 0.01,
    },
    "anual_percentual_finep": {     # FINEP tem uma coluna de valor mínimo
        "cabecalho": 3,
        "colunas": {
            "nome": "N", "taxaCheia": "O",
            "prazoMaximo": "Q", "carenciaMaxima": "R",
            "valorMinimo": "S", "limite": "T",
        },
        "unidade": "anual",
        "escala": 0.01,
    },
}

ABAS = {
    "Linhas Giro Puro": ("mensal_percentual", "giro"),
    "Linhas Investimento": ("mensal_decimal", "investimento"),
    "Linhas Transportes": ("mensal_decimal", "transportes"),
    "Mais Crédito": ("mensal_decimal", "microcredito"),
    "Linhas Fungetur": ("anual_percentual", "fungetur"),
    "Linhas FINEP": ("anual_percentual_finep", "finep"),
    "FCO Empresarial": ("anual_decimal", "fco"),
    "FCO Rural": ("anual_decimal", "fco"),
    "Produtor Empreendedor": ("mensal_decimal", "rural"),
}


def taxa(valor, unidade, escala):
    """Toda taxa carrega a sua unidade — nunca um número solto."""
    if valor is None or isinstance(valor, str):
        return None
    return {"valor": float(valor) * escala, "unidade": unidade, "tipo": "efetiva"}


def numero(valor):
    if valor is None or isinstance(valor, str):
        return None
    return float(valor)


def extrair_linhas(ws, layout):
    """Lê a tabela de linhas até o primeiro buraco.

    Parar no primeiro nome vazio, em vez de varrer uma faixa fixa, é o que
    impede de recolher o cabeçalho de uma segunda tabela mais abaixo — a aba
    'Produtor Empreendedor' tem uma dessas na linha 15, e ela entrava como se
    fosse uma linha de crédito chamada "LINHA".
    """
    cfg = LAYOUTS[layout]
    linhas = []
    for lin in range(cfg["cabecalho"] + 1, cfg["cabecalho"] + 12):
        nome = ws[f"{cfg['colunas']['nome']}{lin}"].value
        if not isinstance(nome, str) or not nome.strip():
            break
        registro = {"nome": nome.strip(), "origemLinha": lin}
        for campo, col in cfg["colunas"].items():
            if campo == "nome":
                continue
            bruto = ws[f"{col}{lin}"].value
            if campo.startswith("taxa"):
                registro[campo] = taxa(bruto, cfg["unidade"], cfg["escala"])
            else:
                registro[campo] = numero(bruto)
        linhas.append(registro)
    return linhas


def extrair_fator_k(wb):
    ws = wb["Fator K - FGI"]
    tabela, repetidos = {}, []
    for lin in range(9, 111):
        prazo, fator = ws[f"F{lin}"].value, ws[f"G{lin}"].value
        if prazo is None or fator is None:
            continue
        chave = str(int(prazo))
        if chave in tabela:
            # XLOOKUP devolve a primeira ocorrência; a segunda nunca é usada.
            repetidos.append({"prazo": int(prazo), "linha": lin,
                              "fatorIgnorado": float(fator),
                              "fatorEmVigor": tabela[chave]})
            continue
        tabela[chave] = float(fator)
    return {"fatores": tabela, "duplicidadesIgnoradas": repetidos}


# A "Tabela de Encargos" não tem estrutura regular que se possa inferir: os
# grupos ocupam faixas de linhas distintas, a unidade da taxa muda de grupo
# para grupo, e o significado das colunas D a F muda junto. Fungetur e FINEP
# usam a coluna E para o indexador, não para uma taxa com bônus; o FCO usa as
# quatro colunas para o par prioritário/não prioritário. Adivinhar isso pela
# forma da célula erra — então o mapa é declarado, e veio da auditoria.
GRUPOS_ENCARGOS = [
    {"grupo": "Linhas para Capital de Giro", "linhas": (5, 9),
     "unidade": "mensal", "formato": "cheia_bonus"},
    {"grupo": "Linhas para Investimentos", "linhas": (12, 17),
     "unidade": "mensal", "formato": "cheia_bonus"},
    {"grupo": "Linhas Especiais para Transportes", "linhas": (20, 24),
     "unidade": "mensal", "formato": "cheia_bonus"},
    {"grupo": "Linhas Especiais de Giro Puro", "linhas": (27, 27),
     "unidade": "mensal", "formato": "cheia_bonus"},
    {"grupo": "Linhas Especiais Investimentos", "linhas": (30, 30),
     "unidade": "mensal", "formato": "cheia_bonus"},
    {"grupo": "Linha Microcrédito", "linhas": (33, 34),
     "unidade": "mensal", "formato": "cheia_bonus"},
    {"grupo": "Linhas Fungetur", "linhas": (37, 40),
     "unidade": "anual", "formato": "cheia_indexador", "indexador": "SELIC"},
    {"grupo": "Linhas FINEP - PORTE I e II", "linhas": (43, 45),
     "unidade": "anual", "formato": "cheia_indexador", "indexador": "TR"},
    {"grupo": "Linhas FINEP - PORTE III", "linhas": (47, 49),
     "unidade": "anual", "formato": "cheia_indexador", "indexador": "TR"},
    {"grupo": "Linhas FCO Empresarial", "linhas": (53, 59),
     "unidade": "anual", "formato": "fco_prioritario"},
    {"grupo": "Linhas FCO Rural e Verde", "linhas": (61, 62),
     "unidade": "anual", "formato": "fco_prioritario"},
]


def prazo_em_meses(valor):
    """A planilha escreve ora 36, ora "Até 36". As duas dizem a mesma coisa."""
    if isinstance(valor, (int, float)):
        return float(valor)
    if isinstance(valor, str):
        achado = re.search(r"\d+", valor)
        if achado:
            return float(achado.group())
    return None


def extrair_encargos(wb):
    """A aba oculta 'Tabela de Encargos' é a tabela oficial, com vigência."""
    ws = wb["Tabela de Encargos"]
    grupos = []
    for cfg in GRUPOS_ENCARGOS:
        u = cfg["unidade"]
        bloco = {"grupo": cfg["grupo"], "unidadeTaxa": u, "linhas": []}
        if "indexador" in cfg:
            bloco["indexador"] = cfg["indexador"]
        primeira, ultima = cfg["linhas"]
        for lin in range(primeira, ultima + 1):
            nome = ws[f"B{lin}"].value
            if not isinstance(nome, str) or not nome.strip():
                continue
            registro = {
                "nome": nome.strip(),
                "origemLinha": lin,
                "prazoMaximo": prazo_em_meses(ws[f"G{lin}"].value),
                "carenciaMaxima": prazo_em_meses(ws[f"H{lin}"].value),
                "limite": numero(ws[f"I{lin}"].value),
            }
            if cfg["formato"] == "cheia_bonus":
                registro["taxaCheia"] = taxa(ws[f"C{lin}"].value, u, 1.0)
                registro["taxaBonus"] = taxa(ws[f"E{lin}"].value, u, 1.0)
            elif cfg["formato"] == "cheia_indexador":
                registro["taxaCheia"] = taxa(ws[f"C{lin}"].value, u, 1.0)
                registro["taxaBonus"] = None
                registro["taxaIndexador"] = taxa(ws[f"E{lin}"].value, u, 1.0)
            else:  # fco_prioritario
                registro["municipioPrioritario"] = {
                    "taxaCheia": taxa(ws[f"C{lin}"].value, u, 1.0),
                    "taxaBonus": taxa(ws[f"D{lin}"].value, u, 1.0),
                }
                registro["municipioNaoPrioritario"] = {
                    "taxaCheia": taxa(ws[f"E{lin}"].value, u, 1.0),
                    "taxaBonus": taxa(ws[f"F{lin}"].value, u, 1.0),
                }
            bloco["linhas"].append(registro)
        grupos.append(bloco)
    return grupos


def principal(caminho, saida):
    wb = openpyxl.load_workbook(caminho, data_only=True)

    parametros = {
        "versao": "2024-12-16",
        "vigenciaInicio": "2024-12-16",
        "origem": caminho.split("/")[-1],
        "observacaoDeUnidade": (
            "Toda taxa traz a unidade em que a planilha a escreveu. A conversão "
            "para o período de cálculo é responsabilidade do motor, e cada aba "
            "converte de um jeito: FCO usa (1+i_anual)^(1/12)-1 e Fungetur e "
            "FINEP usam (1+i_anual)^(22/252)-1."
        ),
        "tabelaDeEncargos": extrair_encargos(wb),
        "fatorKFGI": extrair_fator_k(wb),
        "linhasPorAba": {},
    }

    for aba, (layout, familia) in ABAS.items():
        if aba not in wb.sheetnames:
            continue
        parametros["linhasPorAba"][aba] = {
            "familia": familia,
            "layout": layout,
            "unidadeTaxa": LAYOUTS[layout]["unidade"],
            "linhas": extrair_linhas(wb[aba], layout),
        }

    with open(saida, "w", encoding="utf-8") as fh:
        json.dump(parametros, fh, ensure_ascii=False, indent=2)

    # O mesmo conteúdo sai também como módulo ES, que é o que o motor importa.
    # Importar o .json exigiria atributos de importação, e um PWA que precisa
    # funcionar offline em qualquer navegador não deve depender disso. O JSON
    # continua sendo o artefato de auditoria, legível sem executar nada.
    modulo = saida.replace("dados/PARAMETROS_FINANCEIROS.json", "js/data/parametros.js")
    os.makedirs(os.path.dirname(modulo), exist_ok=True)
    with open(modulo, "w", encoding="utf-8") as fh:
        fh.write(
            "/**\n"
            " * Parâmetros financeiros — GERADO, não editar à mão.\n"
            " *\n"
            " * Sai de `ferramentas/extrair_parametros.py`, que lê a planilha de\n"
            " * referência. Para mudar um valor aqui, muda-se a planilha e roda-se o\n"
            " * extrator de novo; assim a origem de cada número continua rastreável.\n"
            " *\n"
            " * Toda taxa vem com a unidade em que a planilha a escreveu. A conversão\n"
            " * para o período de cálculo é do motor, e cada família de produto\n"
            " * converte de um jeito.\n"
            " */\n\n"
            "export const PARAMETROS = Object.freeze(\n"
        )
        json.dump(parametros, fh, ensure_ascii=False, indent=2)
        fh.write("\n);\n\nexport default PARAMETROS;\n")

    n = sum(len(v["linhas"]) for v in parametros["linhasPorAba"].values())
    print(f"grupos na Tabela de Encargos: {len(parametros['tabelaDeEncargos'])}")
    print(f"linhas por aba: {n}")
    print(f"prazos com fator K: {len(parametros['fatorKFGI']['fatores'])}")
    print(f"duplicidades ignoradas: {parametros['fatorKFGI']['duplicidadesIgnoradas']}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        sys.exit(__doc__)
    principal(sys.argv[1], sys.argv[2])
